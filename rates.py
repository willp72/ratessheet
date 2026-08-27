"""
Savings rate collector.

Pulls best-buy rates from HL, Meteor, Raisin and Flagstone, normalises them
into fixed term buckets, and writes a top 10 per bucket to a Google Sheet.

Run locally:      python rates.py --local
Run in Actions:   python rates.py
"""

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass, asdict

import requests
from bs4 import BeautifulSoup

# Amount held constant so week-on-week comparisons are like for like.
# Products whose published minimum exceeds this are excluded: a rate a
# £1,000 saver cannot actually open is not a rate.
AMOUNT = 1000

# Anna's report order. Keys are internal, values are the labels in the output.
BUCKETS = [
    ("easy_access", "Easy access"),
    ("6m", "6m"),
    ("9m", "9m"),
    ("12m", "12m"),
    ("24m", "24m"),
    ("36m", "36m"),
    ("60m", "60m"),
    ("isa_easy_access", "Easy access ISA"),
    ("isa_12m", "12m ISA"),
    ("isa_24m", "24m ISA"),
]

SOURCES = ["HL", "Meteor", "Flagstone", "Raisin"]

HEALTH: dict[str, object] = {}

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
)


@dataclass
class Product:
    source: str
    bank: str
    rate: float
    term_raw: str
    bucket: str
    isa: bool
    min_deposit: float | None = None    # None means the source does not publish one
    view: str = ""                      # which listing it came from, e.g. "joint"


# ---------------------------------------------------------------------------
# Term normalisation
# ---------------------------------------------------------------------------

def normalise_term(raw: str, isa: bool) -> str | None:
    """Map a source's term label onto one of Anna's buckets. None means drop it."""
    t = raw.lower().strip()

    if any(k in t for k in ("easy access", "instant access", "easy-access")):
        return "isa_easy_access" if isa else "easy_access"

    # Notice accounts are neither easy access nor fixed. Drop them.
    if "notice" in t:
        return None

    months = None
    m = re.search(r"(\d+(?:\.\d+)?)\s*(month|year|yr|m\b|y\b)", t)
    if m:
        n = float(m.group(1))
        unit = m.group(2)
        months = n * 12 if unit.startswith(("year", "yr", "y")) else n

    if months is None:
        return None

    months = int(round(months))
    if isa:
        return {12: "isa_12m", 24: "isa_24m"}.get(months)
    return {6: "6m", 9: "9m", 12: "12m", 24: "24m", 36: "36m", 60: "60m"}.get(months)


def parse_money(txt: str) -> float | None:
    m = re.search(r"£\s*([\d,]+(?:\.\d+)?)", txt)
    return float(m.group(1).replace(",", "")) if m else None


def parse_rate(txt: str) -> float | None:
    m = re.search(r"(\d+\.\d+)\s*%", txt)
    return float(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# Source: Hargreaves Lansdown  (needs a browser: Next.js hydrates the tables)
# ---------------------------------------------------------------------------

HL_GENERAL = "https://www.hl.co.uk/savings/latest-savings-rates-and-products"
HL_ISA = "https://www.hl.co.uk/savings/latest-savings-rates-and-products?filter=cash-isa"


def parse_hl(html: str, isa: bool) -> list[Product]:
    """HL renders one <table> per term with a Bank / AER / Term column set."""
    soup = BeautifulSoup(html, "lxml")
    out = []
    for table in soup.find_all("table"):
        heads = [th.get_text(" ", strip=True).lower() for th in table.select("thead th")]
        if not heads or "bank" not in heads[0]:
            continue
        try:
            i_aer = next(i for i, h in enumerate(heads) if h == "aer")
            i_term = next(i for i, h in enumerate(heads) if h == "term")
        except StopIteration:
            continue
        i_min = next((i for i, h in enumerate(heads) if "min opening" in h), None)

        for tr in table.select("tbody tr"):
            tds = tr.find_all("td")
            if len(tds) <= max(i_aer, i_term):
                continue
            bank = re.sub(r"\s+", " ", tds[0].get_text(" ", strip=True))
            bank = re.sub(r"^Market Leading Rate\s*", "", bank).strip()
            rate = parse_rate(tds[i_aer].get_text())
            term_raw = tds[i_term].get_text(" ", strip=True)
            bucket = normalise_term(term_raw, isa)
            min_dep = (parse_money(tds[i_min].get_text())
                       if i_min is not None and len(tds) > i_min else None)
            if rate and bucket:
                out.append(Product("HL", bank, rate, term_raw, bucket, isa, min_dep))
    return out


# ---------------------------------------------------------------------------
# Source: Meteor  (server rendered, plain fetch is enough)
# ---------------------------------------------------------------------------

# All four URLs from the brief. The filter0 param is applied client-side, so
# these very likely return identical HTML, but fetching all four costs two
# extra plain requests and removes any doubt about ISA-only products being
# absent from the general pages. Dedupe in top_ten() handles the overlap.
METEOR = {
    "easy-access general": "https://savings.meteoram.com/savings/easy-access?filter0=general",
    "fixed-term general": "https://savings.meteoram.com/savings/fixed-term?filter0=general",
    "easy-access isa": "https://savings.meteoram.com/savings/easy-access?filter0=cash-isa",
    "fixed-term isa": "https://savings.meteoram.com/savings/fixed-term?filter0=cash-isa",
}


def _meteor_card(node):
    """Walk up from a rate node to the product card (the bit with the bank name)."""
    card = node
    for _ in range(10):
        card = card.parent
        if card is None:
            return None
        if card.find(["h2", "h3", "h4"]):
            return card
    return None


def meteor_pages(url: str) -> list[str]:
    """
    Fetch every page of a Meteor listing.

    Meteor paginates server-side with &page=N and links each page in a
    .pagination block, so page 1 alone gives about 10 products. The filter0
    param is also server-side, which is why all four brief URLs are fetched.
    """
    first = fetch_plain(url)
    soup = BeautifulSoup(first, "lxml")

    pages = {1}
    for a in soup.select(".pagination a[href]"):
        m = re.search(r"[?&]page=(\d+)", a["href"])
        if m:
            pages.add(int(m.group(1)))

    out = [first]
    for n in sorted(pages)[1:]:
        sep = "&" if "?" in url else "?"
        out.append(fetch_plain(f"{url}{sep}page={n}"))
    return out


def parse_meteor(html: str) -> list[Product]:
    """
    Every Meteor card states its own Type and Term, and carries tick/cross icons
    for General and Cash ISA eligibility. Trust the card, never the page: the
    easy access page carries promoted fixed term cards, and the ?filter0= query
    param is applied client-side so a plain fetch ignores it.
    """
    soup = BeautifulSoup(html, "lxml")
    out = []

    for node in soup.select(".rateInfoValue"):
        rate = parse_rate(re.sub(r"\s+", " ", node.get_text(" ")))
        if not rate:
            continue
        card = _meteor_card(node)
        if card is None:
            continue

        bank = card.find(["h2", "h3", "h4"]).get_text(" ", strip=True)
        text = re.sub(r"\s+", " ", card.get_text(" "))

        m_type = re.search(r"Type\s+(Easy Access|Fixed Term|Notice)", text)
        if not m_type:
            continue
        kind = m_type.group(1)
        if kind == "Notice":
            continue

        if kind == "Easy Access":
            term_raw = "Easy Access"
        else:
            m_term = re.search(
                r"Term\s+(\d+(?:\.\d+)?\s*(?:month|months|year|years))", text
            )
            if not m_term:
                continue
            term_raw = m_term.group(1)

        # Eligibility. Standard cards show both labels with a tick or a cross.
        # Promoted cards show a single label and no icons at all.
        general = isa = False
        icons = card.select("[data-name]")
        if icons:
            for icon in icons:
                label = icon.parent.get_text(" ", strip=True)
                if not label and icon.parent.parent:
                    label = icon.parent.parent.get_text(" ", strip=True)
                ok = icon.get("data-name") == "yes_tick"
                if "Cash ISA" in label:
                    isa = isa or ok
                elif "General" in label:
                    general = general or ok
        else:
            isa = "Cash ISA" in text
            general = "General" in text

        m_dep = re.search(r"Deposit amount\s*(£[\d,]+)", text)
        min_dep = parse_money(m_dep.group(1)) if m_dep else None

        for is_isa in (v for v, on in ((False, general), (True, isa)) if on):
            bucket = normalise_term(term_raw, is_isa)
            if bucket:
                out.append(Product("Meteor", bank, rate, term_raw, bucket, is_isa, min_dep))

    return out


# ---------------------------------------------------------------------------
# Source: Raisin  (needs a browser; sort by term to force all rows to render)
# ---------------------------------------------------------------------------

RAISIN_READY = '[class*="styles-module_rate__"]'
HL_READY = "table tbody tr"

RAISIN_EASY = "https://www.raisin.com/en-gb/savings-accounts/easy-access-savings-accounts/"
RAISIN_FIXED = "https://www.raisin.com/en-gb/savings-accounts/fixed-rate-bonds/"


def parse_raisin(html: str, page_kind: str) -> list[Product]:
    soup = BeautifulSoup(html, "lxml")
    out = []
    for cell in soup.select('[class*="styles-module_rate__"]'):
        row = cell.parent
        text = re.sub(r"\s+", " ", row.get_text(" ")).strip()
        rate = parse_rate(text)
        if not rate:
            continue

        if page_kind == "easy":
            term_raw = "Easy Access"
        else:
            m = re.search(r"AER\s+(.+?)\s+Max", text)
            term_raw = m.group(1) if m else ""
            # Raisin appends badges: "1 Year Sharia account", "6 months Easy access"
            term_raw = re.split(r"\s+(?:Sharia|Easy)\b", term_raw)[0].strip()

        bucket = normalise_term(term_raw, isa=False)
        if not bucket:
            continue

        img = row.find("img", alt=True)
        bank = img["alt"] if img else "unknown"
        out.append(Product("Raisin", bank, rate, term_raw, bucket, False))
    return out


# ---------------------------------------------------------------------------
# Source: Flagstone ISA page  (schema.org JSON in the HTML, plain fetch)
# ---------------------------------------------------------------------------

# Flagstone is the only platform that splits its listing by account type in
# the URL. HL, Meteor and Raisin show one personal table covering both, so
# individual and joint only need handling here.
_FS = "https://clients.direct.flagstoneim.com/build-your-sample-portfolio?accounttype="
FLAGSTONE_SAMPLE_TYPES = {
    "individual": _FS + "individual",
    "joint": _FS + "joint",
}
FLAGSTONE_SAMPLE_READY = "article"

def parse_flagstone_sample(html: str, view: str = "") -> list[Product]:
    """
    The sample portfolio renders one <article> per product:
    "Kent Reliance 4.18% | 4.18% Rate (AER | Gross) Fixed 12 months Term
     £1,000 - £1,000,000 Min - Max deposit Add"
    The first percentage is AER, the second Gross. Take the first.
    """
    soup = BeautifulSoup(html, "lxml")
    out = []
    for art in soup.find_all("article"):
        text = re.sub(r"\s+", " ", art.get_text(" ", strip=True))

        rates = re.findall(r"(\d+\.\d+)%", text)
        if not rates:
            continue
        rate = float(rates[0])

        m = re.search(
            r"(Fixed\s+\d+\s+(?:month|months|year|years)|Instant access)", text, re.I
        )
        if not m:
            continue
        term_raw = m.group(1)

        bank = re.split(r"\s*(?:Best Flagstone|\d+\.\d+%)", text)[0].strip()

        m_dep = re.search(r"(£[\d,]+)\s*[-\u2013]\s*£[\d,]+\s*Min\s*-\s*Max", text)
        min_dep = parse_money(m_dep.group(1)) if m_dep else None

        bucket = normalise_term(term_raw, isa=False)
        if bucket:
            out.append(Product("Flagstone", bank or "unknown", rate, term_raw,
                               bucket, False, min_dep, view))
    return out


FLAGSTONE_PAGES = 5     # ~16 products a page, so ~80. Page 1 is rate-descending.


def fetch_flagstone_pages(url: str, pages: int = FLAGSTONE_PAGES) -> list[str]:
    """
    Return the HTML of the first few pages of the sample portfolio.

    The listing is paginated with numbered buttons, and changing page replaces
    the articles rather than appending to them, so each page has to be captured
    as it is shown. Page 1 alone is not enough here: Flagstone sorts by rate
    descending, and its best rates carry £5,000 minimums that the £1,000 filter
    strips out, leaving its eligible products on later pages.
    """
    from playwright.sync_api import sync_playwright

    out = []
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(user_agent=UA, viewport={"width": 1400, "height": 1200})
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
        dismiss_cookies(page)
        page.wait_for_selector(FLAGSTONE_SAMPLE_READY, timeout=30000)
        out.append(page.content())

        # Click the page numbers by their label. Class names here are generated
        # by styled-components and change on every deploy, so they are useless
        # as selectors; the digits on the buttons are stable.
        for n in range(2, pages + 1):
            try:
                btn = page.get_by_role("button", name=str(n), exact=True)
                if btn.count() == 0:
                    break
                btn.first.scroll_into_view_if_needed(timeout=5000)
                btn.first.click(timeout=5000)
                page.wait_for_timeout(1600)
                page.wait_for_selector(FLAGSTONE_SAMPLE_READY, timeout=15000)
                out.append(page.content())
            except Exception:
                break

        browser.close()

    print(f"    flagstone pages captured: {len(out)}", file=sys.stderr)
    return out


FLAGSTONE = {
    # Public marketing pages, same Umbraco template, same schema.org blocks.
    # The client sample-portfolio tool is deliberately excluded: it needs a
    # login and it is Flagstone's client area.
    # Only the ISA page. The instant-access and fixed-rates pages publish no
    # minimum deposit in their markup, and the sample portfolio covers the same
    # non-ISA products with a verified minimum, so they are redundant here.
    "cash ISA": ("https://www.flagstoneim.com/personal/cash-isa", True),
}


def parse_flagstone(html: str, isa: bool) -> list[Product]:
    out = []
    for blob in re.findall(
        r'<script type="application/ld\+json">(\{.*?)</script>', html, re.S
    ):
        try:
            d = json.loads(blob)
        except json.JSONDecodeError:
            continue
        if d.get("@type") != "FinancialProduct":
            continue
        rate = parse_rate(d.get("annualPercentageRate", ""))
        name = d.get("name", "")
        brand = d.get("brand", name)
        term_raw = name.replace(brand, "").strip()
        bucket = normalise_term(term_raw, isa)
        if rate and bucket:
            out.append(Product("Flagstone", brand, rate, term_raw, bucket, isa))
    return out


# ---------------------------------------------------------------------------
# Fetching
# ---------------------------------------------------------------------------

def fetch_plain(url: str) -> str:
    r = requests.get(url, headers={"User-Agent": UA}, timeout=30)
    r.raise_for_status()
    return r.text


def fetch_rendered(url: str, ready: str, prep=None) -> str:
    """
    Load with a real browser and wait for `ready` to appear.

    Never wait for networkidle on these sites. They run Datadog RUM, Exponea,
    Clarity and Optimizely, which beacon on a timer, so the network never goes
    quiet and the wait burns the full timeout on a page that loaded in seconds.
    Wait for the content instead.
    """
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(user_agent=UA, viewport={"width": 1400, "height": 1200})
        page.goto(url, wait_until="domcontentloaded", timeout=45000)
        dismiss_cookies(page)
        page.wait_for_selector(ready, timeout=30000)
        if prep:
            prep(page)
        html = page.content()
        browser.close()
    return html


def scroll_until_stable(page, selector: str, limit: int = 25) -> int:
    """Scroll until the number of matching elements stops growing."""
    last = -1
    stalls = 0
    for _ in range(limit):
        count = page.locator(selector).count()
        if count == last:
            stalls += 1
            if stalls >= 2:      # two stalls, not one: lazy loads need a beat
                break
        else:
            stalls = 0
        last = count
        page.mouse.wheel(0, 5000)
        page.wait_for_timeout(900)
    return last


def click_load_more(page, button, selector: str, limit: int = 30) -> int:
    """
    Click a 'Load more' button until it disappears.

    Raisin paginates behind a button rather than on scroll, so no amount of
    wheeling will reveal rows 11 onwards.
    """
    for i in range(limit):
        btn = page.locator(button)
        try:
            if btn.count() == 0 or not btn.first.is_visible():
                break
            btn.first.scroll_into_view_if_needed(timeout=5000)
            btn.first.click(timeout=5000)
            page.wait_for_timeout(1200)
        except Exception:
            break
    return page.locator(selector).count()


def dismiss_cookies(page):
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('Accept all')",
        "button:has-text('Accept All')",
        "[data-testid='uc-accept-all-button']",
    ]:
        try:
            page.click(sel, timeout=2500)
            page.wait_for_timeout(600)
            return
        except Exception:
            continue


RAISIN_LOAD_MORE = '[class*="styles-module_loadMore"] button, button:has-text("Load more")'


def raisin_prep(page, warn_on_single_page=True):
    """
    Set the amount, then exhaust the Load more button.

    Raisin shows 10 products and hides the rest behind Load more. Sorting only
    reorders those 10, which is why an earlier run returned five year bonds and
    nothing else. No sort is needed once every row is loaded, we rank ourselves.
    """
    for sel in ("input[inputmode='numeric']", "input[type='text']"):
        try:
            box = page.locator(sel).first
            box.click(timeout=3000)
            box.fill(str(AMOUNT))
            page.keyboard.press("Enter")
            page.wait_for_timeout(2000)
            break
        except Exception:
            continue

    got = click_load_more(page, RAISIN_LOAD_MORE, RAISIN_READY)
    print(f"    raisin rows loaded: {got}", file=sys.stderr)
    if warn_on_single_page and got <= 10:
        print("    WARNING: Load more did not fire, only page 1 captured", file=sys.stderr)


def hl_prep(page):
    got = scroll_until_stable(page, HL_READY)
    print(f"    hl rows after scroll: {got}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Collection and output
# ---------------------------------------------------------------------------

def collect() -> list[Product]:
    products: list[Product] = []

    def attempt(label, fn):
        try:
            got = fn()
            print(f"  {label}: {len(got)} products", file=sys.stderr)
            HEALTH[label] = len(got)
            products.extend(got)
        except Exception as e:
            print(f"  {label}: FAILED {type(e).__name__}: {e}", file=sys.stderr)
            HEALTH[label] = f"FAILED: {type(e).__name__}"

    print("Collecting...", file=sys.stderr)
    attempt("HL general", lambda: parse_hl(fetch_rendered(HL_GENERAL, HL_READY, hl_prep), False))
    attempt("HL ISA", lambda: parse_hl(fetch_rendered(HL_ISA, HL_READY, hl_prep), True))

    for kind, url in METEOR.items():
        attempt(
            f"Meteor {kind}",
            lambda u=url: [p for h in meteor_pages(u) for p in parse_meteor(h)],
        )

    for acct, url in FLAGSTONE_SAMPLE_TYPES.items():
        attempt(
            f"Flagstone {acct}",
            lambda u=url, a=acct: [
                p for h in fetch_flagstone_pages(u) for p in parse_flagstone_sample(h, a)
            ],
        )

    for kind, (url, isa) in FLAGSTONE.items():
        attempt(
            f"Flagstone {kind}",
            lambda u=url, i=isa: parse_flagstone(fetch_plain(u), i),
        )

    attempt("Raisin easy", lambda: parse_raisin(fetch_rendered(RAISIN_EASY, RAISIN_READY,
                                        lambda pg: raisin_prep(pg, warn_on_single_page=False)), "easy"))
    attempt("Raisin fixed", lambda: parse_raisin(fetch_rendered(RAISIN_FIXED, RAISIN_READY, raisin_prep), "fixed"))

    return products


def top_ten(products: list[Product], source: str | None = None) -> dict[str, list[Product]]:
    """
    Rank within each bucket. Duplicates across platforms are kept, as requested,
    but the exact same listing scraped twice from one platform is not a duplicate,
    it's a double count. Those are dropped.

    The dedupe signature deliberately ignores `view`, so a product offered to
    both individual and joint holders takes one slot rather than two, while a
    product only available on one of them still gets in.
    """
    out = {}
    for key, _ in BUCKETS:
        rows, seen = [], set()
        for p in products:
            if p.bucket != key:
                continue
            if source and p.source != source:
                continue
            if p.min_deposit is not None and p.min_deposit > AMOUNT:
                continue
            sig = (p.source, p.bank.lower().strip(), p.rate, p.term_raw.lower())
            if sig in seen:
                continue
            seen.add(sig)
            rows.append(p)
        rows.sort(key=lambda p: -p.rate)
        out[key] = rows[:10]
    return out


def report_rows(ranked: dict[str, list[Product]], title: str) -> list[list]:
    """The ten tables, rates only, in Anna's order. Used by sheet and xlsx."""
    rows = [[title], []]
    for key, label in BUCKETS:
        rows.append([label])
        got = ranked[key]
        for i in range(10):
            rows.append([i + 1, f"{got[i].rate:.2f}" if i < len(got) else ""])
        rows.append([])
    return rows


def render_text(ranked: dict[str, list[Product]]) -> str:
    lines = []
    for key, label in BUCKETS:
        lines.append(label)
        rows = ranked[key]
        for i in range(10):
            if i < len(rows):
                lines.append(f"{i + 1}\t{rows[i].rate:.2f}")
            else:
                lines.append(f"{i + 1}\t")
        lines.append("")
    return "\n".join(lines)


def write_sheet(ranked: dict[str, list[Product]], products: list[Product]):
    """Tab 1: the report as Anna wants it. Tab 2: full audit trail."""
    import gspread
    from datetime import date
    from google.oauth2.service_account import Credentials

    creds_json = os.environ["GOOGLE_CREDENTIALS"]
    sheet_id = os.environ["SHEET_ID"]

    creds = Credentials.from_service_account_info(
        json.loads(creds_json),
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)

    today = date.today().isoformat()

    def tab(title, rows_needed=200):
        existing = [w.title for w in sh.worksheets()]
        return (sh.worksheet(title) if title in existing
                else sh.add_worksheet(title, rows=rows_needed, cols=6))

    # Pooled report across all four platforms.
    rows = report_rows(ranked, f"All platforms, {today}, \u00a3{AMOUNT:,} deposit")
    rows.append(["Source health"])
    for label, count in HEALTH.items():
        rows.append([label, count])

    report = tab("Report")
    report.clear()
    report.update(values=rows, range_name="A1")

    # One tab per platform, same ten tables, that platform only.
    for name in SOURCES:
        per = top_ten(products, source=name)
        ws = tab(name)
        ws.clear()
        ws.update(
            values=report_rows(per, f"{name} only, {today}, \u00a3{AMOUNT:,} deposit"),
            range_name="A1",
        )

    detail_rows = [["date", "bucket", "rate", "bank", "source",
                    "term as listed", "min deposit", "listing"]]
    for key, label in BUCKETS:
        for p in ranked[key]:
            detail_rows.append([
                today, label, p.rate, p.bank, p.source, p.term_raw,
                p.min_deposit if p.min_deposit is not None else "not published",
                p.view or "personal",
            ])

    detail = sh.worksheet("Detail") if "Detail" in [w.title for w in sh.worksheets()] \
        else sh.add_worksheet("Detail", rows=2000, cols=8)
    detail.append_rows(detail_rows[1:], value_input_option="USER_ENTERED")

    print(f"Wrote {len(products)} products to sheet", file=sys.stderr)


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def build_xlsx(ranked, products, path: str):
    """Two sheets: the report as Anna wants it, and the full audit trail."""
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(name="Arial", bold=True)
    plain = Font(name="Arial")

    def write_block(ws, ranked_set, title):
        for row in report_rows(ranked_set, title):
            ws.append(row if row else [None])
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 12
        ws["A1"].font = bold

    ws = wb.active
    ws.title = "Report"
    write_block(ws, ranked, f"All platforms, {date.today():%d %B %Y}, \u00a3{AMOUNT:,} deposit")
    ws.append(["Source health"])
    ws.cell(ws.max_row, 1).font = bold
    for label, count in HEALTH.items():
        ws.append([label, count])

    for name in SOURCES:
        per = wb.create_sheet(name)
        write_block(
            per,
            top_ten(products, source=name),
            f"{name} only, {date.today():%d %B %Y}, \u00a3{AMOUNT:,} deposit",
        )

    det = wb.create_sheet("Detail")
    det.append(["bucket", "rank", "rate", "bank", "source", "term as listed",
                "min deposit", "listing"])
    for c in det[1]:
        c.font = bold
    for key, label in BUCKETS:
        for i, p in enumerate(ranked[key], 1):
            det.append([label, i, float(p.rate), p.bank, p.source, p.term_raw,
                        p.min_deposit if p.min_deposit is not None else "not published",
                        p.view or "personal"])
    for col, w in zip("ABCDEFGH", (18, 6, 8, 38, 12, 20, 14, 12)):
        det.column_dimensions[col].width = w
    det.freeze_panes = "A2"

    labels = {label for _, label in BUCKETS} | {"Source health"}
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for c in row:
                if c.row > 1 and c.column == 1 and c.value in labels:
                    c.font = bold
                elif c.font is not bold:
                    c.font = plain

    wb.save(path)


def send_email(ranked, products):
    """
    Send from Gmail over SMTP. Authenticating as a real Gmail account means the
    mail leaves Google's servers with valid SPF and DKIM, rather than
    originating from a GitHub runner IP, which matters for getting into a
    bank's inbox rather than its quarantine.
    """
    import smtplib
    import ssl
    from datetime import date
    from email.message import EmailMessage

    user = os.environ.get("GMAIL_USER")
    password = os.environ.get("GMAIL_APP_PASSWORD")
    recipients = [a.strip() for a in os.environ.get("REPORT_TO", "").split(",") if a.strip()]

    if not (user and password and recipients):
        print("Email not configured, skipping", file=sys.stderr)
        return

    thin = [label for key, label in BUCKETS if len(ranked[key]) < 10]
    failed = [k for k, v in HEALTH.items() if isinstance(v, str) or v == 0]

    body = [
        f"Savings rates for {date.today():%d %B %Y}, £{AMOUNT:,} deposit.",
        "",
        render_text(ranked).rstrip(),
        "",
    ]
    if thin:
        body.append(f"Fewer than 10 products available: {', '.join(thin)}.")
    if failed:
        body.append(f"Sources that returned nothing: {', '.join(failed)}.")
        body.append("Those tables are drawn from the remaining sources only.")
    body += [
        "",
        f"{len(products)} products collected. Full detail in the attached file.",
        "Rates are AER. Compiled automatically, worth spot checking before use.",
    ]

    msg = EmailMessage()
    msg["Subject"] = f"Savings rates, {date.today():%d %b %Y}"
    msg["From"] = user
    msg["To"] = ", ".join(recipients)
    msg.set_content("\n".join(body))

    path = f"savings-rates-{date.today():%Y-%m-%d}.xlsx"
    build_xlsx(ranked, products, path)
    with open(path, "rb") as fh:
        msg.add_attachment(
            fh.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=path,
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ssl.create_default_context()) as s:
        s.login(user, password)
        s.send_message(msg)

    print(f"Emailed {len(recipients)} recipient(s)", file=sys.stderr)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--local", action="store_true", help="print only, no sheet write, no email")
    args = ap.parse_args()

    products = collect()
    if not products:
        sys.exit("No products collected. Something upstream changed.")

    ranked = top_ten(products)
    print(render_text(ranked))

    thin = [label for key, label in BUCKETS if len(ranked[key]) < 10]
    if thin:
        print(f"\nUnder 10 results: {', '.join(thin)}", file=sys.stderr)

    shown = [p for v in ranked.values() for p in v]
    unverified = sum(1 for p in shown if p.min_deposit is None)
    excluded = sum(1 for p in products
                   if p.min_deposit is not None and p.min_deposit > AMOUNT)
    print(f"Excluded, minimum above £{AMOUNT:,}: {excluded}", file=sys.stderr)
    print(f"Shown without a published minimum: {unverified}", file=sys.stderr)
    HEALTH[f"excluded, min above £{AMOUNT:,}"] = excluded
    HEALTH["shown, minimum not published"] = unverified

    if not args.local:
        write_sheet(ranked, products)
        send_email(ranked, products)


if __name__ == "__main__":
    main()
